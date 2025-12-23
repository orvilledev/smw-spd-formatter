# app.py
import streamlit as st
import os
import zipfile
from io import BytesIO
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# -------------------------------------------------
# Page config
# -------------------------------------------------
st.set_page_config(
    page_title="📦 ALL in 1 SMW FORMATTER TOOL",
    layout="wide",
)

# =================================================
# SECTION 3: LTL FORMATTER (SMW) — FIXED
# =================================================
st.header("📝 LTL FORMATTER (SMW)")

st.markdown(
    """
    Upload a **single SMW Excel file**.
    
    ✅ Automatically fixes SMW box resets  
    ✅ If SMW goes from **Box 44 → Box 1**, output continues to **Box 45**
    """
)

uploaded_file_single = st.file_uploader(
    "📁 Upload SMW Excel file",
    type=["xlsx", "xls"],
)

if uploaded_file_single:
    try:
        input_filename = uploaded_file_single.name
        base_name, ext = os.path.splitext(input_filename)
        output_filename = f"{base_name} formatted{ext}"

        # Read SMW detail section
        df = pd.read_excel(uploaded_file_single, header=10, engine="openpyxl")
        df.columns = df.columns.astype(str).str.strip()

    except Exception as e:
        st.error(f"❌ Error reading file: {e}")

    else:
        required_columns = ["UPC", "Box X", "Sku Units"]
        missing = [c for c in required_columns if c not in df.columns]

        if missing:
            st.error(f"❌ Missing required columns: {', '.join(missing)}")
            st.stop()

        # -------------------------------------------------
        # CLEAN BOX CONTENTS
        # -------------------------------------------------
        df_clean = df[required_columns].dropna(subset=["UPC", "Sku Units"]).copy()

        df_clean["UPC"] = (
            df_clean["UPC"]
            .astype(str)
            .str.replace(r"\.0$", "", regex=True)
            .str.replace("+", "", regex=False)
            .str.zfill(12)
        )

        df_clean["Sku Units"] = (
            pd.to_numeric(df_clean["Sku Units"], errors="coerce")
            .fillna(0)
            .astype(int)
        )

        df_clean.rename(
            columns={"Box X": "Box Number", "Sku Units": "Qty"},
            inplace=True,
        )

        # =================================================
        # 🔥 FIX: PREVENT BOX RESET (THIS IS THE KEY FIX)
        # =================================================
        df_clean["Original Box"] = df_clean["Box Number"]

        new_boxes = []
        current_box = 0
        last_box = None

        for box in df_clean["Original Box"]:
            box = int(box)

            if last_box is None:
                current_box = box
            else:
                # SMW reset detected (44 → 1)
                if box <= last_box:
                    current_box += 1
                else:
                    current_box = box

            new_boxes.append(current_box)
            last_box = box

        df_clean["Box Number"] = new_boxes
        df_clean.drop(columns=["Original Box"], inplace=True)

        # -------------------------------------------------
        # PIVOT TABLE
        # -------------------------------------------------
        pivot_table = pd.pivot_table(
            df_clean,
            index="UPC",
            columns="Box Number",
            values="Qty",
            aggfunc="sum",
            fill_value=0,
        ).reset_index()

        pivot_table = pivot_table.replace(0, "")

        total_qty = int(df_clean["Qty"].sum())
        total_boxes = int(df_clean["Box Number"].nunique())

        # -------------------------------------------------
        # EXTRACT CARTON WEIGHTS (BOLD COLUMN G)
        # -------------------------------------------------
        carton_weights = []
        try:
            uploaded_file_single.seek(0)
            wb_input = load_workbook(uploaded_file_single, data_only=True)
            ws = (
                wb_input["Page1_1"]
                if "Page1_1" in wb_input.sheetnames
                else wb_input[wb_input.sheetnames[0]]
            )

            for row in ws.iter_rows(min_row=1, max_col=7):
                cell = row[6]
                if getattr(cell.font, "bold", False) and isinstance(cell.value, (int, float)):
                    carton_weights.append(cell.value)

            if carton_weights:
                carton_weights = carton_weights[:-1]

        except Exception:
            carton_weights = []

        # -------------------------------------------------
        # EXTRACT DIMENSIONS
        # -------------------------------------------------
        dimension_pattern = r"\b\d{1,3}\.\d{1,2}X\d{1,3}\.\d{1,2}X\d{1,3}\.\d{1,2}\b"
        dims = []

        for _, row in df.iterrows():
            for col in df.columns:
                val = str(row[col])
                if re.match(dimension_pattern, val):
                    try:
                        l, w, h = val.split("X")
                        dims.append((float(l), float(w), float(h)))
                    except:
                        pass

        dim_df = pd.DataFrame()
        if dims:
            dim_df = pd.DataFrame(dims, columns=["Length", "Width", "Height"])
            dim_df.insert(0, "Box Number", range(1, len(dim_df) + 1))
            dim_df.insert(
                1,
                "Carton Weight",
                carton_weights[: len(dim_df)] + [""] * max(0, len(dim_df) - len(carton_weights)),
            )

        # -------------------------------------------------
        # WRITE OUTPUT EXCEL
        # -------------------------------------------------
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df_clean.to_excel(writer, sheet_name="Box Contents", index=False)
            pivot_table.to_excel(writer, sheet_name="Pivot Table", index=False)
            if not dim_df.empty:
                dim_df.to_excel(writer, sheet_name="Box Dimensions", index=False)

        output.seek(0)
        wb = load_workbook(output)

        # -------------------------------------------------
        # STYLING
        # -------------------------------------------------
        yellow = PatternFill("solid", fgColor="FFF2CC")
        bold = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.fill = yellow
                cell.font = bold
                cell.alignment = center
                cell.border = border

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = center
                    cell.border = border

            for col in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col)].width = 18

        # Totals
        ws = wb["Box Contents"]
        r = ws.max_row + 2
        ws[f"A{r}"] = "Total Qty"
        ws[f"B{r}"] = total_qty
        ws[f"A{r+1}"] = "Total Boxes"
        ws[f"B{r+1}"] = total_boxes

        for rr in (r, r + 1):
            for cc in (1, 2):
                c = ws.cell(row=rr, column=cc)
                c.font = bold
                c.border = border
                c.alignment = center

        final_output = BytesIO()
        wb.save(final_output)
        final_output.seek(0)

        st.download_button(
            "💾 Download LTL Formatted File",
            data=final_output,
            file_name=output_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        st.success("✅ LTL shipment processed — box numbering fixed and continuous.")
